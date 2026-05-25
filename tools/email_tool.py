"""
send_email tool.
  send_all_db=False → ONLY rows saved in THIS session (current search)
  send_all_db=True  → ENTIRE database (all rows ever saved)
Uses plain mem dict for reliable cross-thread state.
"""

import os
import json
import smtplib
import tempfile
from datetime import datetime
from email.mime.text      import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base      import MIMEBase
from email                import encoders
from pathlib              import Path

from agents        import function_tool
from utils.logger  import add_log
from utils.db_helper import get_results_since_id


def _build_excel_from_rows(rows: list) -> bytes:
    """Build an in-memory Excel file from given rows only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADERS = ["#", "Query", "Title", "Summary", "URL", "Saved At", "Tags"]
    WIDTHS  = [6, 22, 30, 45, 35, 18, 20]

    wb = Workbook()
    ws = wb.active
    ws.title = "Research Results"

    h_font = Font(name="Calibri", bold=True, color="1A1A2E", size=11)
    h_fill = PatternFill("solid", fgColor="F0F0EB")
    r_font = Font(name="Calibri", color="111111", size=10)
    bdr    = Border(
        bottom=Side(style="thin", color="E0E0D8"),
        right=Side(style="thin",  color="E0E0D8"),
    )

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, 1):
        vals = [i, r["query"], r["title"], r["summary"],
                r.get("url", ""), r["saved_at"], r.get("tags", "")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = r_font
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))
            cell.border = bdr
        ws.row_dimensions[i + 1].height = 50 if len(r["summary"]) > 120 else 28

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    with open(tmp_path, "rb") as f:
        data = f.read()
    Path(tmp_path).unlink(missing_ok=True)
    return data


def make_email_tool(mem: dict):

    @function_tool
    def send_email(
        to_address: str,
        subject:    str,
        body:       str,
    ) -> str:
        """
        Send a research report email via Gmail SMTP.

        Args:
          to_address: recipient email address
          subject:    email subject line
          body:       plain text research summary
        """
        sender   = os.getenv("EMAIL_SENDER",  "").strip()
        password = os.getenv("EMAIL_PASSWORD", "").strip()

        if not sender or not password:
            add_log(mem, "EMAIL", "SKIPPED — credentials missing in .env")
            return json.dumps({"status": "skipped",
                               "message": "Add EMAIL_SENDER and EMAIL_PASSWORD to .env."})

        # ── Choose rows ──────────────────────────────────────────
        since_id   = mem.get("session_start_db_id", 0)
        rows       = get_results_since_id(since_id)
        rows_label = f"Current search only — {len(rows)} new records"
        excel_name = "research_results_current_search.xlsx"
        excel_bytes = _build_excel_from_rows(rows)

        # ── Build HTML email ─────────────────────────────────────
        table_rows_html = ""
        for r in rows:
            url_cell = (
                f'<a href="{r["url"]}" style="color:#1d6fa4">{r["url"][:50]}…</a>'
                if r.get("url") else "—"
            )
            table_rows_html += f"""
            <tr style="border-bottom:1px solid #e0e0d8">
              <td style="padding:6px 10px;color:#888">{r['id']}</td>
              <td style="padding:6px 10px">{r['query']}</td>
              <td style="padding:6px 10px;font-weight:600;color:#1a1a2e">{r['title']}</td>
              <td style="padding:6px 10px;color:#444">{r['summary']}</td>
              <td style="padding:6px 10px;font-size:0.8em">{url_cell}</td>
              <td style="padding:6px 10px;color:#888;font-size:0.8em">{r.get('tags','')}</td>
              <td style="padding:6px 10px;color:#aaa;font-size:0.75em">{r['saved_at']}</td>
            </tr>"""

        scope_badge_color = "#2a9d8f"
        scope_badge_label = "CURRENT SEARCH ONLY"

        html_body = f"""
<html><body style="font-family:Arial,sans-serif;background:#f5f5f0;color:#111;padding:24px;margin:0">
<div style="max-width:900px;margin:auto;background:#fff;border-radius:8px;border:1px solid #e0e0d8;overflow:hidden">
  <div style="background:#1a1a2e;padding:20px 28px;border-bottom:3px solid #e63946">
    <h2 style="color:#fff;margin:0;font-size:1.4rem">⚡ Research Report</h2>
    <p style="color:#aaa;margin:4px 0 0;font-size:0.75rem;font-family:monospace;letter-spacing:2px">
      {datetime.now().strftime('%B %d, %Y  —  %H:%M')}
    </p>
    <span style="display:inline-block;margin-top:8px;background:{scope_badge_color};color:#fff;
                 padding:3px 10px;border-radius:4px;font-size:0.7rem;font-family:monospace;
                 letter-spacing:1.5px">{scope_badge_label} — {len(rows)} records</span>
  </div>
  <div style="padding:20px 28px;border-bottom:1px solid #e0e0d8">
    <p style="white-space:pre-line;line-height:1.8;color:#333;font-size:0.92rem">{body}</p>
  </div>
  <div style="padding:20px 28px">
    <p style="font-family:monospace;font-size:0.65rem;letter-spacing:2px;color:#1a1a2e;
              text-transform:uppercase;margin-bottom:12px">{rows_label}</p>
    <div style="overflow-x:auto">
    <table style="width:100%;border-collapse:collapse;font-size:0.82rem;color:#222">
      <thead><tr style="background:#f0f0eb">
        {''.join(f'<th style="padding:8px 10px;text-align:left;color:#1a1a2e;font-family:monospace;font-size:0.65rem;letter-spacing:1.5px">{h}</th>' for h in ['#','QUERY','TITLE','SUMMARY','URL','TAGS','SAVED AT'])}
      </tr></thead>
      <tbody>{table_rows_html or '<tr><td colspan="7" style="padding:16px;color:#aaa;text-align:center">No records found</td></tr>'}</tbody>
    </table>
    </div>
  </div>
  <div style="background:#f5f5f0;padding:12px 28px;border-top:1px solid #e0e0d8">
    <p style="color:#aaa;font-size:0.7rem;margin:0;font-family:monospace">
      Generated by Multi-Tool Agent — Nexe-Agent Internship
    </p>
  </div>
</div>
</body></html>"""

        try:
            msg = MIMEMultipart("alternative")
            msg["From"]    = sender
            msg["To"]      = to_address
            msg["Subject"] = subject
            msg.attach(MIMEText(body,      "plain"))
            msg.attach(MIMEText(html_body, "html"))

            part = MIMEBase("application",
                            "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(excel_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{excel_name}"')
            msg.attach(part)

            with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=15) as srv:
                srv.login(sender, password)
                srv.send_message(msg)

            mem["email_count"] = mem.get("email_count", 0) + 1
            add_log(mem, "EMAIL", f"Sent → {to_address} | {rows_label} | {excel_name}")

            return json.dumps({
                "status":         "sent",
                "to":             to_address,
                "rows_sent":      len(rows),
                "scope":          "current_search_only",
                "rows_label":     rows_label,
                "excel_attached": excel_name,
                "timestamp":      datetime.now().isoformat(),
            })

        except smtplib.SMTPAuthenticationError:
            add_log(mem, "EMAIL", "AUTH ERROR")
            return json.dumps({"error": "Gmail auth failed. Use App Password.", "status": "error"})
        except Exception as exc:
            add_log(mem, "EMAIL", f"ERROR: {exc}")
            return json.dumps({"error": str(exc), "status": "error"})

    return send_email
