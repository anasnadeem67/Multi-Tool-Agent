"""
Excel read/write helpers using openpyxl.

BUG FIXED:
  Old code:  Border(bottom=Border().bottom.__class__(...))
  Problem:   Border().bottom returns None on fresh Border() object,
             so None.__class__ == NoneType → TypeError.
  Fix:       Import Side directly and use it: Border(bottom=Side(...))
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import EXCEL_PATH, EXCEL_HEADERS, COLUMN_WIDTHS

# ── Style constants ─────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(name="Calibri", bold=True, color="F5C518", size=11)
ROW_FONT    = Font(name="Calibri", color="F0F0F0", size=10)
ALT_FILL    = PatternFill("solid", fgColor="141414")
EVEN_FILL   = PatternFill("solid", fgColor="0F0F0F")

# ✅ CORRECT: use Side() directly — never Border().bottom.__class__
CELL_BORDER = Border(
    bottom=Side(style="thin", color="2A2A2A"),
    right=Side(style="thin",  color="1A1A1A"),
)


def _write_headers(ws) -> None:
    """Write styled header row to worksheet."""
    for col, (header, width) in enumerate(zip(EXCEL_HEADERS, COLUMN_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def init_excel() -> Workbook:
    """Open existing workbook or create a new one with headers."""
    path = Path(EXCEL_PATH)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        if ws.cell(1, 1).value != "#":
            ws.insert_rows(1)
            _write_headers(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Research Results"
        _write_headers(ws)

    wb.save(EXCEL_PATH)
    return wb


def append_to_excel(
    query:   str,
    title:   str,
    summary: str,
    url:     str = "",
    tags:    str = "",
) -> int:
    """
    Append one research result row.
    Returns the human-readable row index (1-based, excluding header).
    """
    wb = init_excel()
    ws = wb.active

    sheet_row = ws.max_row + 1          # actual Excel row number
    row_index = sheet_row - 1           # human-readable index

    fill = ALT_FILL if row_index % 2 == 0 else EVEN_FILL

    values = [
        row_index,
        query,
        title,
        summary,
        url,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        tags,
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=sheet_row, column=col, value=val)
        cell.font      = ROW_FONT
        cell.fill      = fill
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=(col == 4),   # wrap only Summary column
        )
        cell.border = CELL_BORDER

    ws.row_dimensions[sheet_row].height = 50 if len(summary) > 120 else 28
    wb.save(EXCEL_PATH)
    return row_index
